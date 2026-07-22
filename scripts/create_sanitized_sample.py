"""生成可公开提交的脱敏销售出库示例文件。"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bill_splitter.core import patch_formula_caches


ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "sample_data" / "202601销售出库.xlsx"

HEADERS = [
    "销售域名称", "异动别", "单据日期", "日期", "单号", " 2026-01月",
    "客户全称", "客户简称", "品号", "品名", "规格", "单位名称",
    "已出库业务数量", "单价", "未结算原币金额", "备注", "备注",
    "客户全称", "未结算本币金额", "业务员姓名", "销售部门名称", "含税",
    "未结算原币未税金额", "未结算原币税额", "发出商品成本", "商品类型",
]

ROWS = [
    (datetime(2026, 1, 5), "DEMO-26010001", "示例客户甲有限公司", "客户甲", "DEMO-A01", "示例产品A", "蓝色/标准版", 10, 12.5, "按期交付"),
    (datetime(2026, 1, 6), "DEMO-26010002", "示例客户甲有限公司", "客户甲", "DEMO-B02", "示例产品B", "绿色/标准版", 20, 8.0, "按期交付"),
    (datetime(2026, 1, 8), "DEMO-26010003", "示例客户乙有限公司", "客户乙", "DEMO-C03", "示例产品C", "白色/增强版", 5, 30.0, "样品订单"),
]


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"
    sheet.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    caches: dict[str, float] = {}
    for row_index, (date, order_no, customer, short_name, item_no, item_name, spec, quantity, price, note) in enumerate(ROWS, start=2):
        amount = quantity * price
        values = [
            "示例销售域", "销货", date, date, order_no, f"{customer} 2026-01月",
            customer, short_name, item_no, item_name, spec, "PCS", quantity, price,
            f"=M{row_index}*N{row_index}", "", note, customer, amount, "示例业务员",
            "示例销售部", "是", amount, 0, amount * 0.6, "示例商品",
        ]
        sheet.append(values)
        caches[f"O{row_index}"] = amount
    sheet.auto_filter.ref = f"A1:Z{sheet.max_row}"
    sheet.freeze_panes = "A2"
    for column, width in {"A": 12, "C": 12, "D": 12, "E": 18, "F": 34, "G": 26, "H": 14, "I": 16, "J": 18, "K": 22, "L": 12, "M": 14, "N": 12, "O": 16}.items():
        sheet.column_dimensions[column].width = width
    workbook.save(OUTPUT)
    workbook.close()
    patch_formula_caches(OUTPUT, caches)
    print(OUTPUT)


if __name__ == "__main__":
    main()
