from __future__ import annotations

import io
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

import openpyxl
import xlwt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from PIL import Image as PILImage
from app import BillSplitterApp

from bill_splitter.core import (
    BillSplitter,
    HeaderFooterOverrides,
    PauseController,
    ValidationError,
    chinese_upper_rmb,
    detect_source_schema,
    open_reader,
    patch_formula_caches,
    sanitize_filename,
)


class CoreTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.source = self.root / "source.xlsx"
        self.compact_source = self.root / "compact-source.xlsx"
        self.template = self.root / "template.xlsx"
        self.output = self.root / "out"
        self.output.mkdir()
        self._make_source()
        self._make_compact_source()
        self._make_template()
        self.source_xls = self.root / "source.xls"
        self.template_xls = self.root / "template.xls"
        self._make_xls_files()

    def tearDown(self) -> None:
        self.temp.cleanup()

    def _make_source(self) -> None:
        wb = Workbook()
        ws = wb.active
        headers = [
            "销售域名称", "异动别", "单据日期", "日期", "单号", " 2026-01月",
            "客户全称", "客户简称", "品号", "品名", "规格", "单位名称",
            "已出库业务数量", "单价", "未结算原币金额", "备注", "备注",
        ]
        ws.append(headers)
        rows = [
            ("客户甲 2026-01月", "客户甲", "甲简称", 2, 12.5, 25),
            ("客户甲 2026-01月", "客户甲", "甲简称", 3, 10, 30),
            ("客户/乙 2026-01月", "客户/乙", "乙简称", 4, 5, 20),
            (None, None, None, 1, 1, 1),
        ]
        formula_caches = {}
        for index, (key, customer, short, quantity, price, amount) in enumerate(rows, start=2):
            ws.cell(index, 4, f"2026-01-{index:02d}")
            ws.cell(index, 5, f"NO-{index}")
            ws.cell(index, 6, key)
            ws.cell(index, 7, customer)
            ws.cell(index, 8, short)
            ws.cell(index, 9, f"ITEM-{index}")
            ws.cell(index, 10, f"品名-{index}")
            ws.cell(index, 11, f"规格-{index}")
            ws.cell(index, 12, "PCS")
            ws.cell(index, 13, quantity)
            ws.cell(index, 14, price)
            ws.cell(index, 15, f"=M{index}*N{index}")
            formula_caches[f"O{index}"] = amount
            ws.cell(index, 16, "摘要")
            ws.cell(index, 17, "备注")
            ws.cell(index, 13).fill = PatternFill("solid", fgColor="FFFF00")
        wb.save(self.source)
        patch_formula_caches(self.source, formula_caches)

    def _make_template(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "模板"
        for row in range(1, 16):
            for col in range(1, 16):
                ws.cell(row, col, f"T{row}-{col}")
        ws["A1"] = "标题"
        ws["A1"].font = Font(bold=True, size=18)
        ws.merge_cells("A1:L1")
        ws["B6"] = "客户"
        ws["A12"] = "本月合计"
        ws["B12"] = "=J12"
        ws["H12"] = "=SUM(H8:H11)"
        ws["J12"] = "=SUM(J8:J11)"
        ws["L14"] = datetime(2026, 1, 31)
        ws["A15"] = "附注：截止于 2026 年 01月 31日贵公司欠我司款项。"
        ws.merge_cells("A15:L15")
        ws.column_dimensions["A"].width = 18
        ws.row_dimensions[15].height = 30
        ws.print_area = "A1:L15"
        logo_path = self.root / "template-logo.png"
        PILImage.new("RGB", (24, 16), "#0875c1").save(logo_path)
        ws.add_image(XLImage(io.BytesIO(logo_path.read_bytes())), "E1")
        wb.save(self.template)

    def _make_compact_source(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "日期", "单号", " 2026-06月", "客户全称", "客户简称", "品号", "品名",
            "规格", "单位名称", "已出库业务数量", "单价", "未结算原币金额", "备注", "备注",
        ])
        rows = [
            ("紧凑客户甲有限公司", "紧凑甲", 10, 2.5, 25),
            ("紧凑客户甲有限公司", "紧凑甲", 5, 4, 20),
            ("紧凑客户乙有限公司", "紧凑乙", 3, 10, 30),
        ]
        caches = {}
        for row_index, (customer, short_name, quantity, price, amount) in enumerate(rows, start=2):
            sheet.append([
                datetime(2026, 6, row_index), f"COMPACT-{row_index}", f"{customer} 2026-06月",
                customer, short_name, f"ITEM-{row_index}", f"品名-{row_index}", "规格",
                "PCS", quantity, price, f"=J{row_index}*K{row_index}", "摘要", "备注",
            ])
            caches[f"L{row_index}"] = amount
        workbook.save(self.compact_source)
        workbook.close()
        patch_formula_caches(self.compact_source, caches)

    def _make_xls_files(self) -> None:
        source_book = xlwt.Workbook()
        source_sheet = source_book.add_sheet("数据")
        headers = [
            "销售域名称", "异动别", "单据日期", "日期", "单号", " 2026-01月",
            "客户全称", "客户简称", "品号", "品名", "规格", "单位名称",
            "已出库业务数量", "单价", "未结算原币金额", "备注", "备注",
        ]
        for col, header in enumerate(headers):
            source_sheet.write(0, col, header)
        values = [
            ("客户甲 2026-01月", "客户甲", "甲简称"),
            ("客户/乙 2026-01月", "客户/乙", "乙简称"),
        ]
        yellow = xlwt.easyxf("pattern: pattern solid, fore_colour yellow;")
        for row, (key, customer, short) in enumerate(values, start=1):
            row_values = [None] * 17
            row_values[3:17] = [f"2026-01-0{row + 1}", f"NO-{row + 1}", key, customer, short,
                                 f"ITEM-{row + 1}", "品名", "规格", "PCS", row + 1, 5, (row + 1) * 5, "摘要", "备注"]
            for col, value in enumerate(row_values):
                source_sheet.write(row, col, value, yellow if col == 12 else xlwt.Style.default_style)
        source_book.save(str(self.source_xls))

        template_book = xlwt.Workbook()
        template_sheet = template_book.add_sheet("模板")
        title = xlwt.easyxf("font: bold on, height 360; align: horiz center;")
        template_sheet.write_merge(0, 0, 0, 11, "标题", title)
        for row in range(1, 14):
            for col in range(12):
                template_sheet.write(row, col, f"T{row + 1}-{col + 1}")
        template_sheet.write_merge(14, 14, 0, 11, "附注")
        template_sheet.col(0).width = 18 * 256
        template_book.save(str(self.template_xls))

    def test_end_to_end_xlsx(self) -> None:
        progress = []
        result = BillSplitter(
            self.source, self.template, self.output,
            pause=PauseController(), progress=progress.append,
            confirm_overwrite=lambda _: True,
            header_footer=HeaderFooterOverrides(
                company_title="测试科技有限公司",
                bill_year=2027,
                bill_month=3,
                statement_date=date(2027, 3, 31),
            ),
        ).run()
        self.assertEqual((result.total, result.succeeded, result.failed, result.skipped_rows), (2, 2, 0, 1))
        first = self.output / "客户甲 2026-01月.xlsx"
        second = self.output / "客户_乙 2026-01月.xlsx"
        self.assertTrue(first.exists())
        self.assertTrue(second.exists())
        wb = openpyxl.load_workbook(first, data_only=False)
        ws = wb["销售出库明细"]
        self.assertEqual(ws["B6"].value, "客户甲")
        self.assertEqual(ws["A1"].value, "测试科技有限公司")
        self.assertEqual(ws["A5"].value, "2027年03月对账单")
        self.assertEqual(ws["L15"].value, "2027年3月31日")
        self.assertEqual(ws["L15"].number_format, "@")
        self.assertIn("截止于 2027 年 03月 31日", ws["A16"].value)
        self.assertEqual(ws["C7"].value, "T7-5")
        self.assertEqual(ws["D7"].value, "T7-6")
        self.assertEqual(ws["K7"].value, "T7-13")
        self.assertEqual(ws["L7"].value, "T7-14")
        self.assertEqual(ws["A8"].value, "2026-01-02")
        self.assertEqual(ws["B8"].value, "NO-2")
        self.assertEqual(ws["C8"].value, "甲简称")
        self.assertEqual(ws["H8"].value, 2)
        self.assertEqual(ws["J8"].value, "=H8*I8")
        self.assertEqual(ws["J9"].value, "=H9*I9")
        self.assertEqual(ws["H13"].value, "=SUM(H8:H9)")
        self.assertEqual(ws["J13"].value, "=SUM(J8:J9)")
        self.assertEqual(ws["B13"].value, "=J13")
        self.assertIn("A16:L16", {str(r) for r in ws.merged_cells.ranges})
        self.assertEqual(ws.column_dimensions["A"].width, 18)
        self.assertEqual(ws.row_dimensions[16].height, 30)
        self.assertEqual(ws["H8"].fill.fgColor.rgb, "00FFFF00")
        self.assertEqual(ws.max_column, 12)
        self.assertEqual(len(ws._images), 1)
        for row in range(10, 14):
            self.assertEqual(ws.cell(row, 13).style_id, 0)
            self.assertIsNone(ws.cell(row, 13).fill.fill_type)
            self.assertIsNone(ws.cell(row, 13).border.left.style)
            self.assertEqual(ws.cell(row, 14).style_id, 0)
        self.assertEqual(wb.calculation.calcMode, "auto")
        self.assertTrue(wb.calculation.fullCalcOnLoad)
        self.assertTrue(wb.calculation.forceFullCalc)
        wb.close()
        cached_wb = openpyxl.load_workbook(first, data_only=True)
        cached_ws = cached_wb["销售出库明细"]
        self.assertEqual(cached_ws["J8"].value, 25)
        self.assertEqual(cached_ws["J9"].value, 30)
        self.assertEqual(cached_ws["H13"].value, 5)
        self.assertEqual(cached_ws["J13"].value, 55)
        self.assertEqual(cached_ws["B13"].value, 55)
        cached_wb.close()

    def test_existing_files_need_confirmation(self) -> None:
        existing = self.output / "客户甲 2026-01月.xlsx"
        existing.write_bytes(b"keep")
        with self.assertRaisesRegex(ValidationError, "取消了覆盖"):
            BillSplitter(self.source, self.template, self.output, confirm_overwrite=lambda _: False).run()
        self.assertEqual(existing.read_bytes(), b"keep")
        self.assertEqual(len(list(self.output.iterdir())), 1)

    def test_filename_sanitizing_is_stable(self) -> None:
        used: set[str] = set()
        self.assertEqual(sanitize_filename("CON", used), "_CON.xlsx")
        self.assertEqual(sanitize_filename("con", used), "_con (2).xlsx")
        self.assertEqual(sanitize_filename("a/b:*?", used), "a_b___.xlsx")

    def test_chinese_upper_rmb(self) -> None:
        self.assertEqual(chinese_upper_rmb(0), "零元整")
        self.assertEqual(chinese_upper_rmb(55), "伍拾伍元整")
        self.assertEqual(chinese_upper_rmb(1350), "壹仟叁佰伍拾元整")
        self.assertEqual(chinese_upper_rmb("10001.05"), "壹万零壹元零伍分")

    def test_manual_header_input_parsing(self) -> None:
        self.assertEqual(BillSplitterApp._parse_bill_month("2026-01"), (2026, 1))
        self.assertEqual(BillSplitterApp._parse_bill_month("2026年1月"), (2026, 1))
        self.assertEqual(BillSplitterApp._parse_statement_date("2026-01-31"), date(2026, 1, 31))
        self.assertEqual(BillSplitterApp._parse_statement_date("2026年1月31日"), date(2026, 1, 31))
        self.assertEqual(BillSplitterApp._parse_statement_date("2026-06-31"), date(2026, 6, 30))
        self.assertEqual(BillSplitterApp._parse_statement_date("2026-02-30"), date(2026, 2, 28))
        self.assertEqual(BillSplitterApp._parse_statement_date("2028-02-30"), date(2028, 2, 29))
        with self.assertRaises(ValidationError):
            BillSplitterApp._parse_bill_month("2026-13")
        with self.assertRaises(ValidationError):
            BillSplitterApp._parse_statement_date("2026-00-10")

    def test_xls_xlsx_input_combinations(self) -> None:
        combinations = [
            (self.source, self.template_xls),
            (self.source_xls, self.template),
            (self.source_xls, self.template_xls),
        ]
        for index, (source, template) in enumerate(combinations):
            with self.subTest(source=source.suffix, template=template.suffix):
                output = self.root / f"combo-{index}"
                output.mkdir()
                result = BillSplitter(source, template, output, confirm_overwrite=lambda _: True).run()
                self.assertEqual(result.failed, 0, result.errors)
                self.assertEqual(result.succeeded, 2)
                for generated in result.outputs:
                    wb = openpyxl.load_workbook(generated)
                    self.assertIn("销售出库明细", wb.sheetnames)
                    wb.close()

    def test_compact_202606_layout(self) -> None:
        reader = open_reader(self.compact_source, data_only=True)
        try:
            schema = detect_source_schema(reader)
            self.assertEqual(schema.split_key_column, 3)
            self.assertEqual(schema.customer_column, 4)
            self.assertEqual(schema.output_columns, (1, 2, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14))
        finally:
            reader.close()
        output = self.root / "compact-out"
        output.mkdir()
        result = BillSplitter(self.compact_source, self.template, output, confirm_overwrite=lambda _: True).run()
        self.assertEqual((result.total, result.succeeded, result.failed), (2, 2, 0))
        generated = output / "紧凑客户甲有限公司 2026-06月.xlsx"
        workbook = openpyxl.load_workbook(generated, data_only=False)
        sheet = workbook["销售出库明细"]
        self.assertEqual(sheet["A8"].value, datetime(2026, 6, 2))
        self.assertEqual(sheet["B8"].value, "COMPACT-2")
        self.assertEqual(sheet["C8"].value, "紧凑甲")
        self.assertEqual(sheet["J8"].value, "=H8*I8")
        self.assertEqual(sheet["H13"].value, "=SUM(H8:H9)")
        self.assertEqual(sheet["J13"].value, "=SUM(J8:J9)")
        workbook.close()
        cached = openpyxl.load_workbook(generated, data_only=True)
        cached_sheet = cached["销售出库明细"]
        self.assertEqual(cached_sheet["J8"].value, 25)
        self.assertEqual(cached_sheet["H13"].value, 15)
        self.assertEqual(cached_sheet["J13"].value, 45)
        cached.close()


if __name__ == "__main__":
    unittest.main()
