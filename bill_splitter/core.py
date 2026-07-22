from __future__ import annotations

import copy
import io
import os
import re
import tempfile
import threading
import zipfile
import xml.etree.ElementTree as ET
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Callable, Iterable, Protocol

import openpyxl
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.page import PageMargins, PrintOptions
from openpyxl.worksheet.properties import PageSetupProperties

try:
    import xlrd
except ImportError:  # pragma: no cover - 安装包会包含 xlrd
    xlrd = None


SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}
INVALID_FILENAME = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
RESERVED_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}


class ValidationError(ValueError):
    """输入文件或表格结构不符合要求。"""


class PauseController:
    """线程安全的暂停/继续控制器。"""

    def __init__(self) -> None:
        self._running = threading.Event()
        self._running.set()

    @property
    def paused(self) -> bool:
        return not self._running.is_set()

    def pause(self) -> None:
        self._running.clear()

    def resume(self) -> None:
        self._running.set()

    def checkpoint(self) -> None:
        self._running.wait()


@dataclass
class ProgressUpdate:
    phase: str
    message: str
    completed: int = 0
    total: int = 0
    current_group: str = ""


@dataclass
class GroupData:
    key: str
    customer_name: str
    rows: list[int] = field(default_factory=list)
    output_name: str = ""


@dataclass(frozen=True)
class SourceSchema:
    date_column: int
    order_column: int
    split_key_column: int
    customer_column: int
    short_name_column: int
    quantity_column: int
    amount_column: int
    output_columns: tuple[int, ...]


def _header_text(reader: CellReader, column: int) -> str:
    return display_text(reader.value(1, column)).replace(" ", "")


def detect_source_schema(reader: CellReader) -> SourceSchema:
    """通过业务表头识别扩展版和紧凑版销售出库结构。"""
    expected_after_short = (
        "客户简称", "品号", "品名", "规格", "单位名称",
        "已出库业务数量", "单价", "未结算原币金额", "备注", "备注",
    )
    headers = {column: _header_text(reader, column) for column in range(1, reader.max_column + 1)}
    for short_column in range(5, reader.max_column - 8):
        actual = tuple(headers.get(short_column + offset, "") for offset in range(10))
        unit_matches = actual[4] in {"单位名称", "单位"}
        amount_matches = actual[7] in {"未结算原币金额", "金额"}
        if (
            actual[:4] == expected_after_short[:4]
            and unit_matches
            and actual[5:7] == expected_after_short[5:7]
            and amount_matches
            and actual[8:] == expected_after_short[8:]
            and headers.get(short_column - 1) == "客户全称"
            and headers.get(short_column - 4) in {"日期", "单据日期"}
            and headers.get(short_column - 3) == "单号"
        ):
            return SourceSchema(
                date_column=short_column - 4,
                order_column=short_column - 3,
                split_key_column=short_column - 2,
                customer_column=short_column - 1,
                short_name_column=short_column,
                quantity_column=short_column + 5,
                amount_column=short_column + 7,
                output_columns=(short_column - 4, short_column - 3, *range(short_column, short_column + 10)),
            )
    readable_headers = "、".join(value for value in headers.values() if value)[:240]
    raise ValidationError(
        "无法识别销售出库文件结构。需要包含连续字段：日期、单号、拆分键、客户全称、"
        "客户简称、品号、品名、规格、单位名称、已出库业务数量、单价、未结算原币金额、备注、备注。"
        f"\n当前表头：{readable_headers}"
    )


@dataclass
class SplitResult:
    total: int
    succeeded: int
    failed: int
    skipped_rows: int
    output_dir: Path
    errors: list[str] = field(default_factory=list)
    outputs: list[Path] = field(default_factory=list)


@dataclass(frozen=True)
class HeaderFooterOverrides:
    company_title: str
    bill_year: int
    bill_month: int
    statement_date: date

    def __post_init__(self) -> None:
        if not self.company_title.strip():
            raise ValidationError("公司抬头不能为空。")
        if not 1 <= self.bill_month <= 12:
            raise ValidationError("账单月份必须在 1–12 之间。")


class CellReader(Protocol):
    max_row: int
    max_column: int

    def value(self, row: int, column: int) -> Any: ...
    def copy_cell_to(self, row: int, column: int, target: Any) -> None: ...
    def row_height(self, row: int) -> float | None: ...
    def close(self) -> None: ...


def _first_visible_nonempty_xlsx(workbook: Any) -> Any:
    for sheet in workbook.worksheets:
        if sheet.sheet_state == "visible" and sheet.max_row > 0 and sheet.max_column > 0:
            if any(cell.value is not None for row in sheet.iter_rows() for cell in row):
                return sheet
    raise ValidationError("工作簿中没有可见且非空的工作表。")


class XlsxReader:
    def __init__(self, path: Path, *, data_only: bool = True) -> None:
        try:
            self.workbook = openpyxl.load_workbook(path, data_only=data_only)
            self.sheet = _first_visible_nonempty_xlsx(self.workbook)
        except ValidationError:
            raise
        except Exception as exc:
            raise ValidationError(f"无法读取 Excel 文件“{path.name}”：{exc}") from exc
        self.max_row = self.sheet.max_row
        self.max_column = self.sheet.max_column

    def value(self, row: int, column: int) -> Any:
        return self.sheet.cell(row, column).value

    def copy_cell_to(self, row: int, column: int, target: Any) -> None:
        source = self.sheet.cell(row, column)
        value = source.value
        if isinstance(value, str) and value.startswith("="):
            try:
                value = Translator(value, origin=source.coordinate).translate_formula(target.coordinate)
            except Exception:
                pass
        target.value = value
        apply_style(target, capture_style(source))

    def row_height(self, row: int) -> float | None:
        return self.sheet.row_dimensions[row].height

    def close(self) -> None:
        self.workbook.close()


def _xls_color(book: Any, color_index: int | None, default: str = "000000") -> str:
    if color_index is None:
        return default
    rgb = book.colour_map.get(color_index)
    if not rgb:
        return default
    return "FF%02X%02X%02X" % rgb


def _xls_border_style(value: int) -> str | None:
    return {
        0: None, 1: "thin", 2: "medium", 3: "dashed", 4: "dotted",
        5: "thick", 6: "double", 7: "hair", 8: "mediumDashed",
        9: "dashDot", 10: "mediumDashDot", 11: "dashDotDot",
        12: "mediumDashDotDot", 13: "slantDashDot",
    }.get(value, "thin")


class XlsReader:
    def __init__(self, path: Path) -> None:
        if xlrd is None:
            raise ValidationError("缺少读取 .xls 文件所需的 xlrd 组件。")
        try:
            self.workbook = xlrd.open_workbook(path, formatting_info=True)
            candidates = [s for s in self.workbook.sheets() if getattr(s, "visibility", 0) == 0]
            self.sheet = next((s for s in candidates if s.nrows and s.ncols), None)
            if self.sheet is None:
                raise ValidationError("工作簿中没有可见且非空的工作表。")
        except ValidationError:
            raise
        except Exception as exc:
            raise ValidationError(f"无法读取 Excel 文件“{path.name}”：{exc}") from exc
        self.max_row = self.sheet.nrows
        self.max_column = self.sheet.ncols

    def _cell(self, row: int, column: int) -> Any:
        return self.sheet.cell(row - 1, column - 1)

    def value(self, row: int, column: int) -> Any:
        cell = self._cell(row, column)
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate_as_datetime(cell.value, self.workbook.datemode)
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return None
        return cell.value

    def copy_cell_to(self, row: int, column: int, target: Any) -> None:
        cell = self._cell(row, column)
        target.value = self.value(row, column)
        if cell.xf_index is None or cell.xf_index >= len(self.workbook.xf_list):
            return
        xf = self.workbook.xf_list[cell.xf_index]
        font = self.workbook.font_list[xf.font_index]
        target.font = Font(
            name=font.name or "宋体", size=(font.height / 20) if font.height else 11,
            bold=bool(font.bold), italic=bool(font.italic),
            underline="single" if font.underline_type else None,
            color=_xls_color(self.workbook, font.colour_index),
        )
        pattern = xf.background
        fill_type = "solid" if pattern.fill_pattern else None
        target.fill = PatternFill(fill_type=fill_type, fgColor=_xls_color(self.workbook, pattern.pattern_colour_index, "FFFFFF"))
        borders = xf.border
        target.border = Border(
            left=Side(style=_xls_border_style(borders.left_line_style), color=_xls_color(self.workbook, borders.left_colour_index)),
            right=Side(style=_xls_border_style(borders.right_line_style), color=_xls_color(self.workbook, borders.right_colour_index)),
            top=Side(style=_xls_border_style(borders.top_line_style), color=_xls_color(self.workbook, borders.top_colour_index)),
            bottom=Side(style=_xls_border_style(borders.bottom_line_style), color=_xls_color(self.workbook, borders.bottom_colour_index)),
        )
        align = xf.alignment
        target.alignment = Alignment(
            horizontal={0: None, 1: "left", 2: "center", 3: "right", 4: "fill", 5: "justify"}.get(align.hor_align),
            vertical={0: None, 1: "top", 2: "center", 3: "bottom", 4: "justify"}.get(align.vert_align),
            wrap_text=bool(align.text_wrapped), text_rotation=align.rotation,
        )
        fmt = self.workbook.format_map.get(xf.format_key)
        if fmt and fmt.format_str:
            target.number_format = fmt.format_str

    def row_height(self, row: int) -> float | None:
        info = self.sheet.rowinfo_map.get(row - 1)
        return (info.height / 20) if info and info.height else None

    def close(self) -> None:
        self.workbook.release_resources()


def open_reader(path: Path, *, data_only: bool = True) -> CellReader:
    if path.suffix.lower() == ".xlsx":
        return XlsxReader(path, data_only=data_only)
    if path.suffix.lower() == ".xls":
        return XlsReader(path)
    raise ValidationError(f"不支持的文件格式：{path.suffix or '无扩展名'}")


@dataclass
class TemplateCell:
    row: int
    column: int
    value: Any
    style: dict[str, Any]


@dataclass
class TemplateImage:
    data: bytes
    width: float
    height: float
    anchor: Any


def capture_style(cell: Any) -> dict[str, Any]:
    return {
        "font": copy.copy(cell.font),
        "fill": copy.copy(cell.fill),
        "border": copy.copy(cell.border),
        "alignment": copy.copy(cell.alignment),
        "number_format": cell.number_format,
        "protection": copy.copy(cell.protection),
    }


def apply_style(cell: Any, style: dict[str, Any]) -> None:
    cell.font = copy.copy(style["font"])
    cell.fill = copy.copy(style["fill"])
    cell.border = copy.copy(style["border"])
    cell.alignment = copy.copy(style["alignment"])
    cell.number_format = style["number_format"]
    cell.protection = copy.copy(style["protection"])


class TemplateSnapshot:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.cells: list[TemplateCell] = []
        self.images: list[TemplateImage] = []
        self.row_heights: dict[int, float] = {}
        self.column_widths: dict[int, float] = {}
        self.column_hidden: dict[int, bool] = {}
        self.merges: list[tuple[int, int, int, int]] = []
        self.page_setup: Any = None
        self.page_margins: Any = None
        self.print_options: Any = None
        self.sheet_properties: Any = None
        self.sheet_format: Any = None
        self.sheet_view: Any = None
        self.freeze_panes: Any = None
        self.print_area: str | None = None
        if path.suffix.lower() == ".xlsx":
            self._read_xlsx(path)
        else:
            self._read_xls(path)

    def _read_xlsx(self, path: Path) -> None:
        try:
            wb = openpyxl.load_workbook(path, data_only=False)
            ws = _first_visible_nonempty_xlsx(wb)
        except ValidationError:
            raise
        except Exception as exc:
            raise ValidationError(f"无法读取模板“{path.name}”：{exc}") from exc
        if ws.max_row < 15:
            wb.close()
            raise ValidationError("表头表尾模板必须至少包含 15 行。")
        # 表头会删除 C/D，因此需读取到原 N 列，左移后正好成为输出 K/L。
        # 表尾是在删除 C/D 后插入，VBA 参考结果只保留 A–L。
        for row in list(range(1, 8)) + list(range(9, 16)):
            max_col = 14 if row <= 7 else 12
            for column in range(1, max_col + 1):
                cell = ws.cell(row, column)
                if isinstance(cell, MergedCell):
                    continue
                self.cells.append(TemplateCell(row, column, cell.value, capture_style(cell)))
            dim = ws.row_dimensions[row]
            if dim.height is not None:
                self.row_heights[row] = dim.height
        for column in range(1, 13):
            dim = ws.column_dimensions[get_column_letter(column)]
            if dim.width is not None:
                self.column_widths[column] = dim.width
            self.column_hidden[column] = bool(dim.hidden)
        for merged in ws.merged_cells.ranges:
            min_col, min_row, max_col_range, max_row = range_boundaries(str(merged))
            if max_row <= 7 or min_row >= 9:
                self.merges.append((min_col, min_row, max_col_range, max_row))
        for image in ws._images:
            self.images.append(
                TemplateImage(
                    data=image._data(),
                    width=image.width,
                    height=image.height,
                    anchor=copy.deepcopy(image.anchor),
                )
            )
        self.page_setup = copy.copy(ws.page_setup)
        self.page_margins = copy.copy(ws.page_margins)
        self.print_options = copy.copy(ws.print_options)
        self.sheet_properties = copy.copy(ws.sheet_properties)
        self.sheet_format = copy.copy(ws.sheet_format)
        self.sheet_view = copy.copy(ws.sheet_view)
        self.freeze_panes = ws.freeze_panes
        try:
            self.print_area = str(ws.print_area) if ws.print_area else None
        except Exception:
            self.print_area = None
        wb.close()

    def _read_xls(self, path: Path) -> None:
        reader = XlsReader(path)
        try:
            if reader.max_row < 15:
                raise ValidationError("表头表尾模板必须至少包含 15 行。")
            holder_book = Workbook()
            holder_sheet = holder_book.active
            for row in list(range(1, 8)) + list(range(9, 16)):
                max_col = min(14 if row <= 7 else 12, reader.max_column)
                for column in range(1, max_col + 1):
                    holder = holder_sheet.cell(row, column)
                    reader.copy_cell_to(row, column, holder)
                    self.cells.append(TemplateCell(row, column, holder.value, capture_style(holder)))
                height = reader.row_height(row)
                if height is not None:
                    self.row_heights[row] = height
            for column in range(1, 13):
                info = reader.sheet.colinfo_map.get(column - 1)
                if info:
                    self.column_widths[column] = info.width / 256
                    self.column_hidden[column] = bool(info.hidden)
            for row_lo, row_hi, col_lo, col_hi in reader.sheet.merged_cells:
                min_row, max_row = row_lo + 1, row_hi
                if max_row <= 7 or min_row >= 9:
                    self.merges.append((col_lo + 1, min_row, col_hi, max_row))
        finally:
            reader.close()

    @staticmethod
    def _translated_value(value: Any, origin: str, target: str) -> Any:
        if isinstance(value, str) and value.startswith("="):
            try:
                return Translator(value, origin=origin).translate_formula(target)
            except Exception:
                return value
        return value

    def create_sheet(self, data_rows: int) -> tuple[Workbook, Any, int, int]:
        wb = Workbook()
        ws = wb.active
        ws.title = "销售出库明细"
        last_data_row = 7 + data_rows
        footer_start = last_data_row + 1
        footer_shift = footer_start - 9
        for spec in self.cells:
            target_row = spec.row if spec.row <= 7 else spec.row + footer_shift
            # 旧宏先插入表头与数据，再删除整张表的 C、D 列；表尾是在删除后插入的。
            # 因此只有表头 1–7 行需要跳过 C/D，并将 E 及之后的列左移两列。
            if spec.row <= 7 and spec.column in (3, 4):
                continue
            target_column = spec.column - 2 if spec.row <= 7 and spec.column >= 5 else spec.column
            target = ws.cell(target_row, target_column)
            origin_coord = f"{get_column_letter(spec.column)}{spec.row}"
            target_coord = f"{get_column_letter(target_column)}{target_row}"
            target.value = self._translated_value(spec.value, origin_coord, target_coord)
            apply_style(target, spec.style)
        for row, height in self.row_heights.items():
            target_row = row if row <= 7 else row + footer_shift
            ws.row_dimensions[target_row].height = height
        for column, width in self.column_widths.items():
            letter = get_column_letter(column)
            ws.column_dimensions[letter].width = width
            ws.column_dimensions[letter].hidden = self.column_hidden.get(column, False)
        for min_col, min_row, max_col, max_row in self.merges:
            if min_row >= 9:
                min_row += footer_shift
                max_row += footer_shift
            elif min_row <= 7:
                if min_col >= 5:
                    min_col -= 2
                    max_col -= 2
                elif max_col >= 3:
                    removed = max(0, min(max_col, 4) - max(min_col, 3) + 1)
                    max_col = max(min_col, max_col - removed)
            ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        for image_spec in self.images:
            image = XLImage(io.BytesIO(image_spec.data))
            image.width = image_spec.width
            image.height = image_spec.height
            image.anchor = copy.deepcopy(image_spec.anchor)
            ws.add_image(image)
        # VBA 原逻辑会把表头第 1–5 行统一合并为 A:L。
        for row in range(1, 6):
            for merged in list(ws.merged_cells.ranges):
                if merged.min_row == row and merged.max_row == row:
                    ws.unmerge_cells(str(merged))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        if self.page_setup is not None:
            ws.page_setup = copy.copy(self.page_setup)
            ws.page_margins = copy.copy(self.page_margins)
            ws.print_options = copy.copy(self.print_options)
            ws.sheet_properties = copy.copy(self.sheet_properties)
            ws.sheet_format = copy.copy(self.sheet_format)
            if self.sheet_view is not None:
                ws.sheet_view.showGridLines = self.sheet_view.showGridLines
                ws.sheet_view.showRowColHeaders = self.sheet_view.showRowColHeaders
                ws.sheet_view.zoomScale = self.sheet_view.zoomScale
                ws.sheet_view.zoomScaleNormal = self.sheet_view.zoomScaleNormal
            ws.freeze_panes = self.freeze_panes
        # 模板的打印区域通常是固定 1:15；输出时延伸到实际表尾。
        if self.print_area:
            try:
                min_col, _, max_col, _ = range_boundaries(self.print_area.replace("'", "").split("!")[-1])
                ws.print_area = f"{get_column_letter(min_col)}1:{get_column_letter(max_col)}{footer_start + 6}"
            except Exception:
                ws.print_area = f"A1:L{footer_start + 6}"
        return wb, ws, last_data_row, footer_start


def _decimal_number(value: Any) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        if isinstance(value, str):
            value = value.strip().replace(",", "")
            if not value:
                return None
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _excel_number(value: Decimal) -> int | float:
    return int(value) if value == value.to_integral_value() else float(value)


def _chinese_group(number: int) -> str:
    digits = "零壹贰叁肆伍陆柒捌玖"
    units = ("仟", "佰", "拾", "")
    values = (number // 1000, number // 100 % 10, number // 10 % 10, number % 10)
    result: list[str] = []
    pending_zero = False
    for digit, unit in zip(values, units):
        if digit:
            if pending_zero and result:
                result.append("零")
            result.append(digits[digit] + unit)
            pending_zero = False
        elif result:
            pending_zero = True
    return "".join(result)


def chinese_upper_rmb(value: Decimal | int | float) -> str:
    amount = Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    prefix = "负" if amount < 0 else ""
    amount = abs(amount)
    fen_total = int(amount * 100)
    integer, fraction = divmod(fen_total, 100)
    jiao, fen = divmod(fraction, 10)
    digits = "零壹贰叁肆伍陆柒捌玖"
    if integer == 0:
        integer_text = "零"
    else:
        groups: list[int] = []
        remaining = integer
        while remaining:
            groups.append(remaining % 10000)
            remaining //= 10000
        big_units = ("", "万", "亿", "兆")
        parts: list[str] = []
        zero_pending = False
        for index in range(len(groups) - 1, -1, -1):
            group = groups[index]
            if not group:
                if parts and any(groups[:index]):
                    zero_pending = True
                continue
            if parts and (zero_pending or group < 1000):
                parts.append("零")
            parts.append(_chinese_group(group) + big_units[index])
            zero_pending = False
        integer_text = "".join(parts)
    result = prefix + integer_text + "元"
    if jiao == 0 and fen == 0:
        return result + "整"
    if jiao:
        result += digits[jiao] + "角"
    elif fen:
        result += "零"
    if fen:
        result += digits[fen] + "分"
    return result


def _formula_cache_text(value: Any) -> tuple[str, bool]:
    if isinstance(value, bool):
        return ("1" if value else "0"), False
    if isinstance(value, Decimal):
        return format(value, "f"), False
    if isinstance(value, int):
        return str(value), False
    if isinstance(value, float):
        return (str(int(value)) if value.is_integer() else repr(value)), False
    return str(value), True


def patch_formula_caches(path: Path, values: dict[str, Any]) -> None:
    """为公式单元格写入 Excel 缓存值，使未重算的查看器也能显示结果。"""
    if not values:
        return
    namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    fd, patched_name = tempfile.mkstemp(prefix=".账单缓存_", suffix=".xlsx", dir=path.parent)
    os.close(fd)
    patched = Path(patched_name)
    try:
        with zipfile.ZipFile(path, "r") as source_zip, zipfile.ZipFile(patched, "w") as target_zip:
            for item in source_zip.infolist():
                data = source_zip.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    for _, pair in ET.iterparse(io.BytesIO(data), events=("start-ns",)):
                        prefix, uri = pair
                        try:
                            ET.register_namespace(prefix, uri)
                        except ValueError:
                            pass
                    root = ET.fromstring(data)
                    cells = {cell.attrib.get("r", "").upper(): cell for cell in root.iter(f"{{{namespace}}}c")}
                    for coordinate, cached_value in values.items():
                        cell = cells.get(coordinate.upper())
                        if cell is None or cell.find(f"{{{namespace}}}f") is None:
                            continue
                        text, is_string = _formula_cache_text(cached_value)
                        value_node = cell.find(f"{{{namespace}}}v")
                        if value_node is None:
                            value_node = ET.SubElement(cell, f"{{{namespace}}}v")
                        value_node.text = text
                        if is_string:
                            cell.set("t", "str")
                        elif cell.attrib.get("t") in {"str", "inlineStr"}:
                            cell.attrib.pop("t", None)
                    data = ET.tostring(root, encoding="utf-8", xml_declaration=False)
                target_zip.writestr(item, data)
        os.replace(patched, path)
    finally:
        if patched.exists():
            try:
                patched.unlink()
            except OSError:
                pass


def display_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sanitize_filename(name: str, used: set[str], max_length: int = 160) -> str:
    clean = INVALID_FILENAME.sub("_", name).strip().rstrip(". ")
    clean = re.sub(r"\s+", " ", clean) or "未命名账单"
    if clean.split(".")[0].upper() in RESERVED_NAMES:
        clean = "_" + clean
    clean = clean[:max_length].rstrip(". ") or "未命名账单"
    base = clean
    index = 2
    while clean.casefold() in used:
        suffix = f" ({index})"
        clean = base[: max_length - len(suffix)].rstrip(". ") + suffix
        index += 1
    used.add(clean.casefold())
    return clean + ".xlsx"


class BillSplitter:
    def __init__(
        self,
        source_path: str | Path,
        template_path: str | Path,
        output_dir: str | Path,
        *,
        pause: PauseController | None = None,
        progress: Callable[[ProgressUpdate], None] | None = None,
        confirm_overwrite: Callable[[list[Path]], bool] | None = None,
        header_footer: HeaderFooterOverrides | None = None,
    ) -> None:
        self.source_path = Path(source_path).expanduser().resolve()
        self.template_path = Path(template_path).expanduser().resolve()
        self.output_dir = Path(output_dir).expanduser().resolve()
        self.pause = pause or PauseController()
        self.progress = progress or (lambda _: None)
        self.confirm_overwrite = confirm_overwrite or (lambda _: False)
        self.header_footer = header_footer

    def validate_paths(self) -> None:
        for label, path in (("待拆分文件", self.source_path), ("表头表尾模板", self.template_path)):
            if not path.is_file():
                raise ValidationError(f"{label}不存在：{path}")
            if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                raise ValidationError(f"{label}只支持 .xlsx 或 .xls 格式。")
        if self.source_path == self.template_path:
            raise ValidationError("待拆分文件和表头表尾模板不能是同一个文件。")
        if not self.output_dir.is_dir():
            raise ValidationError(f"输出目录不存在：{self.output_dir}")
        try:
            with tempfile.NamedTemporaryFile(dir=self.output_dir, prefix=".账单工具_", delete=True):
                pass
        except OSError as exc:
            raise ValidationError(f"输出目录不可写：{exc}") from exc

    def _emit(self, phase: str, message: str, completed: int = 0, total: int = 0, current: str = "") -> None:
        self.progress(ProgressUpdate(phase, message, completed, total, current))

    def _group_rows(self, source: CellReader, schema: SourceSchema | None = None) -> tuple[list[GroupData], int]:
        schema = schema or detect_source_schema(source)
        if source.max_row < 2:
            raise ValidationError("源工作表没有可拆分的数据行。")
        month_label = display_text(source.value(1, schema.split_key_column))
        groups: OrderedDict[str, GroupData] = OrderedDict()
        skipped = 0
        for row in range(2, source.max_row + 1):
            if row % 50 == 0:
                self.pause.checkpoint()
                self._emit("读取", f"正在读取并分组：第 {row}/{source.max_row} 行")
            key = display_text(source.value(row, schema.split_key_column))
            customer = display_text(source.value(row, schema.customer_column))
            if not key and customer and month_label:
                key = f"{customer}{month_label}"
            if not key:
                skipped += 1
                continue
            normalized = key.casefold()
            group = groups.get(normalized)
            if group is None:
                group = GroupData(key=key, customer_name=customer)
                groups[normalized] = group
            elif not group.customer_name and customer:
                group.customer_name = customer
            group.rows.append(row)
        if not groups:
            raise ValidationError("F 列没有可用的拆分键，且无法用 G 列客户名称与 F1 月份生成拆分键。")
        used: set[str] = set()
        result = list(groups.values())
        for group in result:
            group.output_name = sanitize_filename(group.key, used)
        return result, skipped

    def _write_group(
        self,
        source: CellReader,
        template: TemplateSnapshot,
        group: GroupData,
        destination: Path,
        formula_source: CellReader | None = None,
        schema: SourceSchema | None = None,
    ) -> None:
        schema = schema or detect_source_schema(source)
        wb, ws, last_data_row, footer_start = template.create_sheet(len(group.rows))
        temp_path: Path | None = None
        try:
            ws["B6"] = group.customer_name
            if self.header_footer is not None:
                settings = self.header_footer
                ws["A1"] = settings.company_title.strip()
                ws["A5"] = f"{settings.bill_year}年{settings.bill_month:02d}月对账单"
                date_row = footer_start + 5
                date_cell = ws.cell(date_row, 12)
                # 直接写文本，避免部分 WPS/预览器把 Excel 日期序列显示成 46203。
                date_cell.value = (
                    f"{settings.statement_date.year}年"
                    f"{settings.statement_date.month}月{settings.statement_date.day}日"
                )
                date_cell.number_format = "@"
                note_cell = ws.cell(footer_start + 6, 1)
                if isinstance(note_cell.value, str):
                    deadline = (
                        f"截止于 {settings.statement_date.year} 年 "
                        f"{settings.statement_date.month:02d}月 {settings.statement_date.day:02d}日"
                    )
                    note_cell.value = re.sub(
                        r"截止于\s*\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日",
                        deadline,
                        note_cell.value,
                        count=1,
                    )
            formula_reader = formula_source or source
            formula_caches: dict[str, Any] = {}
            quantity_total = Decimal("0")
            amount_total = Decimal("0")
            for output_offset, source_row in enumerate(group.rows):
                self.pause.checkpoint()
                output_row = 8 + output_offset
                for output_col, source_col in enumerate(schema.output_columns, start=1):
                    target_cell = ws.cell(output_row, output_col)
                    formula_reader.copy_cell_to(source_row, source_col, target_cell)
                    if isinstance(target_cell.value, str) and target_cell.value.startswith("="):
                        cached_value = source.value(source_row, source_col)
                        if cached_value is not None:
                            formula_caches[target_cell.coordinate] = cached_value
                quantity = _decimal_number(source.value(source_row, schema.quantity_column))
                amount = _decimal_number(source.value(source_row, schema.amount_column))
                if quantity is not None:
                    quantity_total += quantity
                if amount is not None:
                    amount_total += amount
                height = source.row_height(source_row)
                if height is not None:
                    ws.row_dimensions[output_row].height = height
            sum_row = footer_start + 3
            ws.cell(sum_row, 8).value = f"=SUM(H8:H{last_data_row})"
            ws.cell(sum_row, 10).value = f"=SUM(J8:J{last_data_row})"
            formula_caches[ws.cell(sum_row, 8).coordinate] = _excel_number(quantity_total)
            formula_caches[ws.cell(sum_row, 10).coordinate] = _excel_number(amount_total)
            total_text_cell = ws.cell(sum_row, 2)
            if isinstance(total_text_cell.value, str) and total_text_cell.value.startswith("="):
                normalized_formula = total_text_cell.value.replace("$", "").replace(" ", "").upper()
                if normalized_formula == f"=J{sum_row}":
                    formula_caches[total_text_cell.coordinate] = _excel_number(amount_total)
                elif "[DBNUM2]" in normalized_formula:
                    formula_caches[total_text_cell.coordinate] = chinese_upper_rmb(amount_total)
            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            fd, temp_name = tempfile.mkstemp(prefix=".账单工具_", suffix=".xlsx", dir=self.output_dir)
            os.close(fd)
            temp_path = Path(temp_name)
            wb.save(temp_path)
            patch_formula_caches(temp_path, formula_caches)
            os.replace(temp_path, destination)
            temp_path = None
        finally:
            wb.close()
            if temp_path and temp_path.exists():
                try:
                    temp_path.unlink()
                except OSError:
                    pass

    def run(self) -> SplitResult:
        self.validate_paths()
        self._emit("读取", "正在读取销售出库文件……")
        source = open_reader(self.source_path, data_only=True)
        formula_source: CellReader | None = None
        try:
            if self.source_path.suffix.lower() == ".xlsx":
                formula_source = open_reader(self.source_path, data_only=False)
            else:
                formula_source = source
            schema = detect_source_schema(source)
            groups, skipped = self._group_rows(source, schema)
            self._emit("模板", "正在读取表头表尾模板……", 0, len(groups))
            template = TemplateSnapshot(self.template_path)
            conflicts = [self.output_dir / group.output_name for group in groups if (self.output_dir / group.output_name).exists()]
            if conflicts and not self.confirm_overwrite(conflicts):
                raise ValidationError("用户取消了覆盖已有文件，未开始生成账单。")
            result = SplitResult(len(groups), 0, 0, skipped, self.output_dir)
            for index, group in enumerate(groups, start=1):
                self.pause.checkpoint()
                self._emit("写入", f"正在生成：{group.key}", index - 1, len(groups), group.key)
                destination = self.output_dir / group.output_name
                try:
                    self._write_group(
                        source,
                        template,
                        group,
                        destination,
                        formula_source=formula_source,
                        schema=schema,
                    )
                    result.succeeded += 1
                    result.outputs.append(destination)
                except Exception as exc:
                    result.failed += 1
                    result.errors.append(f"{group.key}：{exc}")
                self._emit("写入", f"已完成 {index}/{len(groups)}", index, len(groups), group.key)
            self._emit("完成", "账单拆分完成。", len(groups), len(groups))
            return result
        finally:
            if formula_source is not None and formula_source is not source:
                formula_source.close()
            source.close()
